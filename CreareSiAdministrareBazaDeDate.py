import openpyxl

class AdministrareBazaDate:
#Scrie in baza de date-Parametrii: linia,coloana unde sa scrie, valuarea de scris, calea spre baza de date
    def Scriere_In_Baza_Date(self,lin,col,valuare, cale):
        bazaDate = openpyxl.load_workbook(cale)
        foaie = bazaDate.active
        celula = foaie.cell(row=lin, column=col)
        celula.value = valuare
        bazaDate.save(cale)
#Citeste din baza de date-Parametrii: linia,coloana de unde sa citeasca, calea spre baza de date
    def Citire_Din_Baza_Date(self,lin,col,cale):
        bazaDate = openpyxl.load_workbook(cale)
        foaie = bazaDate.active
        celula = foaie.cell(row=lin, column=col)
        valuare = celula.value
        return valuare
#Returneaza linia si coloana maxima din baza de date-Parametrii: calea spre baza de date
    def Dimensiune_Baza_Date(self,cale):
        bazaDate = openpyxl.load_workbook(cale)
        foaie = bazaDate.active
        return([foaie.max_row,foaie.max_column])
#Sterge linia corespunzatoare ID-ului din baza de date-Parametrii: ID-ul obiectului ce trebuie sters, calea spre baza de date
    def Stergere_Din_Baza_Date(self,ID,cale):
        bazaDate = AdministrareBazaDate()
        dimensiune = bazaDate.Dimensiune_Baza_Date(cale)
        if ID + 1 == dimensiune[0]:
            for i in range(dimensiune[1]):
                bazaDate.Scriere_In_Baza_Date(dimensiune[0],i+1,None,cale)
        else:
            for i in range(dimensiune[1]):
                bazaDate.Scriere_In_Baza_Date(ID + 1,i+1,bazaDate.Citire_Din_Baza_Date(ID+2,i+1,cale),cale)
                bazaDate.Scriere_In_Baza_Date(dimensiune[0],i+1,None,cale)
